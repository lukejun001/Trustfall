#!/usr/bin/env python3
"""Build a visual Excel summary of the Trustfall final-pilot statistics.

Pulls live operational counts from production, the frozen snapshot from the
training manifest, and the final metrics from the evaluation comparison, then
renders styled tables and charts suited to a quick on-camera walkthrough.
Uses only aggregate counts and metrics -- no message bodies or secrets.
"""
import argparse
import json
import os
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import psycopg
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "0B1220"; CARD = "111E3D"; CYAN = "22D3EE"; BLUE = "2563EB"
GREEN = "34D399"; AMBER = "FBBF24"; VIOLET = "8B5CF6"; SLATE = "334155"
WHITE = "F8FAFC"; PALE = "EAF2FF"; INK = "111827"; GREY = "E2E8F0"

TITLE = Font(name="Calibri", size=22, bold=True, color=WHITE)
SUB = Font(name="Calibri", size=11, color="CBD5E1")
H = Font(name="Calibri", size=13, bold=True, color=WHITE)
KPI = Font(name="Calibri", size=30, bold=True, color=WHITE)
KPILBL = Font(name="Calibri", size=10, color="CBD5E1")
TH = Font(name="Calibri", size=11, bold=True, color=WHITE)
TD = Font(name="Calibri", size=11, color=INK)
TDB = Font(name="Calibri", size=11, bold=True, color=INK)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color=GREY)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(color):
    return PatternFill("solid", fgColor=color)


def fetch_live(url):
    with psycopg.connect(url) as c, c.cursor() as cur:
        cur.execute('SELECT count(*) FROM "CollectedMessage" WHERE "isSyntheticTestFixture"=false'); real = cur.fetchone()[0]
        cur.execute('SELECT count(*) FROM "MessageLabel"'); labels = cur.fetchone()[0]
        cur.execute('SELECT count(*) FROM "LabelAssignment"'); assigned = cur.fetchone()[0]
        cur.execute('SELECT count(*) FROM "ExpertAdjudication"'); expert = cur.fetchone()[0]
        cur.execute('SELECT "candidateReason", count(*) FROM "ExpertAdjudication" GROUP BY 1'); reasons = dict(cur.fetchall())
        cur.execute('SELECT count(*) FROM "TeracParticipant" WHERE wave=%s AND status=%s', ("label", "label_completed")); completed = cur.fetchone()[0]
        cur.execute('''SELECT count(l.id) FROM "CollectedMessage" m LEFT JOIN "MessageLabel" l ON l."messageId"=m.id
                       WHERE m."isSyntheticTestFixture"=false GROUP BY m.id'''); dist = Counter(r[0] for r in cur.fetchall())
    return {"real": real, "labels": labels, "assigned": assigned, "expert": expert, "reasons": reasons,
            "completed": completed, "dist": dict(sorted(dist.items()))}


def style_band(ws, row, last_col, color, height=None):
    for col in range(1, last_col + 1):
        ws.cell(row=row, column=col).fill = fill(color)
    if height:
        ws.row_dimensions[row].height = height


def kpi_block(ws, row, col, value, label, accent):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = KPI; cell.alignment = CENTER; cell.fill = fill(CARD)
    lbl = ws.cell(row=row + 1, column=col, value=label)
    lbl.font = KPILBL; lbl.alignment = CENTER; lbl.fill = fill(CARD)
    bar = ws.cell(row=row + 2, column=col); bar.fill = fill(accent)
    ws.row_dimensions[row + 2].height = 4


def header_row(ws, row, headers, start=1, color=BLUE):
    for i, text in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=text)
        c.font = TH; c.fill = fill(color); c.alignment = CENTER; c.border = BORDER


def data_row(ws, row, values, start=1, bold_first=True, zebra=False):
    for i, value in enumerate(values):
        c = ws.cell(row=row, column=start + i, value=value)
        c.font = TDB if (bold_first and i == 0) else TD
        c.alignment = LEFT if i == 0 else CENTER
        c.border = BORDER
        if zebra:
            c.fill = fill(PALE)


def title_banner(ws, title, subtitle, last_col, generated):
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=last_col)
    t = ws.cell(row=1, column=1, value=title); t.font = TITLE; t.alignment = LEFT
    style_band(ws, 1, last_col, NAVY); style_band(ws, 2, last_col, NAVY)
    ws.row_dimensions[1].height = 30; ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    s = ws.cell(row=3, column=1, value=f"{subtitle}    ·    {generated}"); s.font = SUB; s.alignment = LEFT
    style_band(ws, 3, last_col, NAVY); ws.row_dimensions[3].height = 18


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--database-url", default=os.getenv("DATABASE_URL"))
    parser.add_argument("--stamp", default=(Path("tmp/pilot_stamp.txt").read_text().strip() if Path("tmp/pilot_stamp.txt").exists() else None))
    parser.add_argument("--output", default=str(Path.home() / "Desktop" / "Trustfall_Pilot_Stats.xlsx"))
    args = parser.parse_args()
    if not args.database_url:
        raise SystemExit("Set DATABASE_URL or pass --database-url.")
    if not args.stamp:
        raise SystemExit("Provide --stamp (frozen snapshot timestamp).")

    live = fetch_live(args.database_url)
    manifest = json.loads(Path(f"wave2_training_data/{args.stamp}/manifest.json").read_text())
    comparison = json.loads(Path(f"wave2_evaluation/{args.stamp}/comparison.json").read_text())
    generated = "Generated " + datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    prov = manifest["provenance_counts"]; splits = manifest["split_counts"]; excluded = manifest.get("excluded", {})
    base, fine = comparison["baseline"], comparison["fine_tuned"]

    wb = Workbook()

    # ---------------- Sheet 1: Summary ----------------
    ws = wb.active; ws.title = "Summary"; ws.sheet_view.showGridLines = False
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15.5
    title_banner(ws, "TRUSTFALL — Final Pilot Statistics",
                 "Privacy-first human data → Qwen LoRA scam-safety model", 8, generated)

    ws.cell(row=5, column=1, value="LIVE OPERATIONAL  (production, changes as the pilot runs)").font = Font(bold=True, size=12, color=BLUE)
    kpi_block(ws, 6, 1, live["real"], "real sanitized messages", CYAN)
    kpi_block(ws, 6, 3, live["labels"], "human labels saved", BLUE)
    kpi_block(ws, 6, 5, live["expert"], "expert adjudications", VIOLET)
    kpi_block(ws, 6, 7, live["completed"], "labelers completed", GREEN)

    ws.cell(row=10, column=1, value="FROZEN FINAL-PILOT SNAPSHOT  (immutable once built)").font = Font(bold=True, size=12, color=BLUE)
    kpi_block(ws, 11, 1, manifest["eligible_messages"], "final-target messages", CYAN)
    kpi_block(ws, 11, 3, splits["train"], "training examples", BLUE)
    kpi_block(ws, 11, 5, splits["validation"], "validation examples", AMBER)
    kpi_block(ws, 11, 7, splits["test"], "test holdout (used once)", GREEN)

    ws.cell(row=15, column=1, value="MODEL — base Qwen vs. LoRA on the frozen test split").font = Font(bold=True, size=12, color=BLUE)
    kpi_block(ws, 16, 1, f"{base['valid_json_rate']*100:.0f}% → {fine['valid_json_rate']*100:.0f}%", "valid JSON", CYAN)
    kpi_block(ws, 16, 3, f"{base['risk_exact_accuracy']*100:.0f}% → {fine['risk_exact_accuracy']*100:.0f}%", "risk-level exact", BLUE)
    kpi_block(ws, 16, 5, f"{base['risk_within_one_accuracy']*100:.0f}% → {fine['risk_within_one_accuracy']*100:.0f}%", "risk within one level", AMBER)
    kpi_block(ws, 16, 7, f"{base['scam_type_exact_accuracy']*100:.0f}% → {fine['scam_type_exact_accuracy']*100:.0f}%", "scam-type exact", GREEN)

    note = ws.cell(row=20, column=1,
                   value="⚠ Small-data limitation: " + manifest.get("small_data_limitation", ""))
    note.font = Font(italic=True, size=10, color="92400E"); note.alignment = LEFT
    ws.merge_cells(start_row=20, start_column=1, end_row=21, end_column=8)
    style_band(ws, 20, 8, "FEF3C7"); style_band(ws, 21, 8, "FEF3C7")

    # ---------------- Sheet 2: Data & Provenance ----------------
    ws2 = wb.create_sheet("Data & Provenance"); ws2.sheet_view.showGridLines = False
    for col, w in {"A": 26, "B": 14, "C": 14, "D": 14}.items():
        ws2.column_dimensions[col].width = w
    title_banner(ws2, "Data quality & provenance", "How messages become final targets", 4, generated)

    # Label distribution table + chart
    ws2.cell(row=5, column=1, value="Messages by number of worker labels (live)").font = H
    style_band(ws2, 5, 4, NAVY)
    header_row(ws2, 6, ["Labels per message", "Messages"], 1, BLUE)
    dist = live["dist"]; r = 7
    for k in range(0, 6):
        data_row(ws2, r, [f"{k} label{'s' if k != 1 else ''}", dist.get(k, 0)], 1, zebra=(r % 2 == 1)); r += 1
    dist_end = r - 1
    chart = BarChart(); chart.type = "col"; chart.title = "Label coverage per message"
    chart.height = 7.5; chart.width = 13; chart.legend = None
    data = Reference(ws2, min_col=2, min_row=6, max_row=dist_end)
    cats = Reference(ws2, min_col=1, min_row=7, max_row=dist_end)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    chart.y_axis.title = "Messages"; ws2.add_chart(chart, "F5")

    # Provenance table + pie
    pr = dist_end + 2
    ws2.cell(row=pr, column=1, value="Final-target provenance (frozen)").font = H
    style_band(ws2, pr, 4, NAVY)
    header_row(ws2, pr + 1, ["Source", "Messages"], 1, BLUE)
    data_row(ws2, pr + 2, ["Worker consensus", prov.get("worker_consensus", 0)], 1, zebra=True)
    data_row(ws2, pr + 3, ["Expert adjudication", prov.get("expert_adjudication", 0)], 1)
    data_row(ws2, pr + 4, ["Total eligible", manifest["eligible_messages"]], 1, zebra=True)
    pie = PieChart(); pie.title = "Provenance of final targets"; pie.height = 7.5; pie.width = 11
    pdata = Reference(ws2, min_col=2, min_row=pr + 1, max_row=pr + 3)
    pcats = Reference(ws2, min_col=1, min_row=pr + 2, max_row=pr + 3)
    pie.add_data(pdata, titles_from_data=True); pie.set_categories(pcats)
    ws2.add_chart(pie, "F22")

    # Adjudication reasons + exclusions
    ar = pr + 6
    ws2.cell(row=ar, column=1, value="Expert adjudication candidates & exclusions").font = H
    style_band(ws2, ar, 4, NAVY)
    header_row(ws2, ar + 1, ["Category", "Count"], 1, BLUE)
    reasons = live["reasons"]
    rows = [("Single worker label", reasons.get("single_label", 0)),
            ("No worker labels", reasons.get("no_labels", 0)),
            ("Low agreement", reasons.get("low_agreement", 0)),
            ("Excluded: not reviewable", excluded.get("not_reviewable", 0)),
            ("Excluded: insufficient labels", excluded.get("not_enough_labels", 0))]
    for i, (lbl, val) in enumerate(rows):
        data_row(ws2, ar + 2 + i, [lbl, val], 1, zebra=(i % 2 == 0))

    # ---------------- Sheet 3: Model Results ----------------
    ws3 = wb.create_sheet("Model Results"); ws3.sheet_view.showGridLines = False
    for col, w in {"A": 30, "B": 16, "C": 16}.items():
        ws3.column_dimensions[col].width = w
    title_banner(ws3, "Base Qwen vs. fine-tuned LoRA", f"One-time evaluation on the {comparison['test_examples']}-message frozen test split", 3, generated)
    ws3.cell(row=5, column=1, value="Higher is better · percentages").font = SUB
    header_row(ws3, 6, ["Metric", "Baseline", "Fine-tuned"], 1, BLUE)
    metrics = [("Valid JSON rate", base["valid_json_rate"], fine["valid_json_rate"]),
               ("Risk-level exact accuracy", base["risk_exact_accuracy"], fine["risk_exact_accuracy"]),
               ("Risk within-one-level accuracy", base["risk_within_one_accuracy"], fine["risk_within_one_accuracy"]),
               ("Scam-type exact accuracy", base["scam_type_exact_accuracy"], fine["scam_type_exact_accuracy"]),
               ("Red-flag F1", base["red_flag_f1"], fine["red_flag_f1"])]
    for i, (lbl, b, f) in enumerate(metrics):
        row = 7 + i
        data_row(ws3, row, [lbl], 1)
        for col, val in ((2, b), (3, f)):
            c = ws3.cell(row=row, column=col, value=round(val * 100, 1)); c.font = TD
            c.alignment = CENTER; c.border = BORDER; c.number_format = '0.0"%"'
            if row % 2 == 1:
                c.fill = fill(PALE)
    end = 6 + len(metrics)
    chart = BarChart(); chart.type = "col"; chart.grouping = "clustered"
    chart.title = "Baseline vs. fine-tuned (frozen test split)"; chart.height = 9; chart.width = 18
    chart.y_axis.title = "Percent"; chart.y_axis.scaling.min = 0; chart.y_axis.scaling.max = 100
    data = Reference(ws3, min_col=2, max_col=3, min_row=6, max_row=end)
    cats = Reference(ws3, min_col=1, min_row=7, max_row=end)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    ws3.add_chart(chart, "E6")
    n = ws3.cell(row=end + 2, column=1, value="⚠ " + manifest.get("small_data_limitation", ""))
    n.font = Font(italic=True, size=10, color="92400E"); n.alignment = LEFT
    ws3.merge_cells(start_row=end + 2, start_column=1, end_row=end + 3, end_column=3)
    style_band(ws3, end + 2, 3, "FEF3C7"); style_band(ws3, end + 3, 3, "FEF3C7")

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
